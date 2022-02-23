# -*- coding: utf-8 -*-

import cfg
import telebot
from time import sleep
from sys import argv
import requests
from bs4 import BeautifulSoup
import json
from telebot import types
import sqlite3
from multiprocessing import Process
import openpyxl
import random
import datetime

bot = telebot.TeleBot(cfg.token)
script, link, max_ads, need, min_year, status_message, message_text, id, country, id_parser = argv

token_pt = {'authorization': 'Bearer 43f550c77c013b4213abb4320181ddadeed95dfc'}
token_pl = {'authorization': 'Bearer 0693cab5648f869a1a06b45a52ec262b4938db7d'}
token_ro = {'authorization': 'Bearer 75d09c5a8eac171fd10b046fffe0c22ff2a3d4c8'}
token_kz = {'authorization': 'Bearer 684db05d3e6b005c85357737665a5c2c7d4ce752'}
token_bg = {'authorization': 'Bearer a91b7e736f9f6966b73d7d45ad485a94732a036d'}
token_ua = {'authorization': 'Bearer b13cce4ac41e7cf96890d64c6b300b6c0f1302d7'}


proxies1 = {
    'http': 'http://S6yTc5:YJacDv@195.158.195.186:8000',
    'https': 'http://S6yTc5:YJacDv@195.158.195.186:8000',
}
proxies2 = {
    'http': 'http://S6yTc5:YJacDv@195.158.195.244:8000',
    'https': 'http://S6yTc5:YJacDv@195.158.195.244:8000',
}





def parse_tr(link, max_ads, need, min_year, status_message, message_text, chat_id, country, thread, id_parser, time_now):
    checks = []
    if country == 'pt':
        token = token_pt
    elif country == 'pl':
        token = token_pl
    elif country == 'ro':
        token = token_ro
    elif country == 'kz':
        token = token_kz
    elif country == 'bg':
        token = token_bg
    elif country == 'ua':
        token = token_ua
    main_link = link
    count_need = 0
    bad_number1 = 0
    bad_count1 = 0
    bad_year1 = 0
    bad_link_checked1 = 0
    all_checked = 0
    for start in range(int(thread), 101, 15):
        rand = random.randint(1, 2)
        if rand == 1:
            proxies = proxies1
        else:
            proxies = proxies2
        try:
            if int(count_need) >= int(need):
                break
            main_link = main_link.replace('https://www.olx.'+str(country)+'/', 'https://m.olx.'+str(country)+'/')
            if '?' in main_link:
                check_page = requests.get(str(main_link) + '&page='+str(start))
            else:
                check_page = requests.get(str(main_link) + '?page=' + str(start))
            checked = check_page.text
            soup = BeautifulSoup(checked, "html.parser")
            link_ad_1 = soup.find_all(class_="detailsLink")
            for i in range(len(link_ad_1)-1):
                conn = sqlite3.connect("parser.db")
                c = conn.cursor()
                count_need = 0
                c.execute("SELECT id_user FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                for id_user, in c.fetchall():
                    count_need += 1
                conn.close()
                if int(count_need) >= int(need):
                    break
                link_ad_2 = link_ad_1[i].get("href")
                link_ad = link_ad_2[:link_ad_2.find("html")+4]
                if len(link_ad) < 20:
                    continue
                if link_ad in checks:
                    continue
                if "www.olx."+str(country)+"/" in link_ad:
                    pass
                else:
                    continue
                try:
                    all_checked += 1
                    checks.append(link_ad)
                    check_page = requests.get(link_ad)
                    checked = check_page.text
                    soup = BeautifulSoup(checked, "html.parser")
                    gorod_1 = checked.find('"cityName"') + 12
                    gorod_2 = checked.find('"', gorod_1)
                    gorod = checked[gorod_1:gorod_2]
                    title = soup.title.string
                    title_1 = title.find('•')
                    title = title[:title_1]
                    name = soup.find(class_="css-owpmn2-Text eu5v0x0").string
                    reg = soup.find(class_="css-1bafgv4-Text eu5v0x0").string
                    created_1 = checked.find('"created":"')+11
                    created_2 = checked.find('+', created_1)
                    created = checked[created_1:created_2].replace('T', ' ')
                    if country == 'kz' or country == 'bg' or country == 'ua':
                        reg = reg[-7:-3]
                    else:
                        reg = reg[-4:]
                    if int(reg) < int(min_year):
                        bad_year1 += 1
                        continue
                    prof_link = soup.find(class_="css-1qj8w5r").get('href')
                    prof_link = 'https://www.olx.' + str(country) + str(prof_link)
                    check_page1 = requests.get(prof_link)
                    checked1 = check_page1.text
                    soup1 = BeautifulSoup(checked1, "html.parser")
                    count_1 = soup1.find_all(class_="lheight22 margintop5")
                    count = len(count_1)
                    if int(count) > int(max_ads):
                        bad_count1 += 1
                        continue
                    id_1 = soup.find(class_="css-sc0ups-BaseStyles").get('href')
                    id = id_1[(id_1.find('id')) + 3:]
                    price_1 = str(soup.find(class_="css-8kqr5l-Text eu5v0x0"))
                    price = str(price_1[(price_1.find('">')) + 2:price_1.find('<!')])
                    if country == "pt":
                        price += ' €'
                    elif country == "pl":
                        price += ' zł'
                    elif country == "ro":
                        price += ' lei / €'
                    elif country == "kz":
                        price += ' ₸'
                    elif country == "bg":
                        price += ' лв'
                    elif country == "ua":
                        price += ' UAH'
                    check_page = requests.get("https://www.olx."+str(country)+"/api/v1/offers/" + str(id) + "/phones/", headers=token, proxies=proxies)
                    checked = check_page.text
                    print(checked)
                    try:
                        number = str(json.loads(checked)['data']['phones'][0].replace(' ', '').replace('-', '')).replace('(', '').replace(')', '')
                        if country == 'pt':
                            number = '351' + str(number)
                        elif country == 'pl':
                            number = '48' + str(number)
                        elif country == 'ro':
                            if int(number[0]) == 0:
                                number = '4' + str(number)
                            else:
                                number = '40' + str(number)
                        elif country == 'kz':
                            if int(number[0]) == 8:
                                number = '7' + str(number[1:])
                            else:
                                number = '7' + str(number)
                        elif country == 'bg':
                            if int(number[0]) == 0:
                                number = '359' + str(number[1:])
                            else:
                                number = '359' + str(number)
                        elif country == 'ua':
                            if int(number[0]) == 0:
                                number = '38' + str(number)
                            else:
                                number = '380' + str(number)

                    except Exception as e:
                        print(e)
                        number = ''
                    if len(str(number)) < 5:
                        bad_number1 += 1
                        continue
                    number = number.replace('+', '')
                    conn = sqlite3.connect("parser.db")
                    c = conn.cursor()
                    a = 0
                    count_views = 0
                    c.execute("SELECT id_user FROM parsed_url WHERE number = '" + str(number) + "';")
                    for id_user, in c.fetchall():
                        if int(id_user) == int(chat_id):
                            a = 1
                            break
                        count_views+=1
                    conn.close()
                    if a == 1:
                        bad_link_checked1 += 1
                        continue
                    img = soup.find(class_="css-1bmvjcs").get('src')
                    link = link_ad
                    conn = sqlite3.connect("parser.db")
                    c = conn.cursor()
                    a = 0
                    c.execute("SELECT id_user FROM parsed_url WHERE url = '" + str(link_ad) + "';")
                    for id_user, in c.fetchall():
                        if int(id_user) == int(chat_id):
                            a = 1
                            break
                    conn.close()
                    if a == 1:
                        continue
                    whatKey = types.InlineKeyboardMarkup()
                    key_last = types.InlineKeyboardButton(text="❌ Закрыть", callback_data="close")
                    whatKey.row(key_last)
                    if status_message == 'ON':
                        message_send = message_text.replace('[title]', title).replace('[link]', link_ad)
                    else:
                        message_send = ''
                    conn = sqlite3.connect("parser.db")
                    c = conn.cursor()
                    c.execute("INSERT INTO parsed_url VALUES(?,?,?,?,?,?,?,?,?,?,?, ?)", (chat_id, id_parser, link, title, name, price, number, count, '0', created, img, gorod))
                    conn.commit()
                    conn.close()
                    bot.send_photo(chat_id, requests.get(img).content,
                                       caption=f"Название: <b>{title}</b>\n"
                                               f"Цена: <b>{price}</b>\n"
                                               f"Имя: <b>{name}</b>\n"
                                               f"Зарегистрирован: <b>{created}</b>\n"
                                               f"Местоположение: <b>{gorod}</b>\n"
                                               f"Номер: <b>+{number}</b>\n\n"
                                               f'Объявление: <a href="'+str(link_ad)+'">*Клик*</a>\n'
                                               f'WhatsApp: <a href="https://api.whatsapp.com/send?phone={number}&text={message_send}">*Клик*</a>\n'
                                               f'Viber: <a href="https://msng.link/o/?{number}=vi">*Клик*</a>\n\n'
                                               f"Видело пользователей: <b>{count_views}</b>",
                                       parse_mode="html",
                                       reply_markup=whatKey)
                    conn = sqlite3.connect("parser.db")
                    c = conn.cursor()
                    count_need = 0
                    c.execute("SELECT id_user FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                    for id_user, in c.fetchall():
                        count_need += 1
                    conn.close()
                    if int(count_need) >= int(need):
                        break
                except Exception as e:
                    pass
        except Exception as e:
            bot.send_message(chat_id, "⚠️ Ошибка")
            break
    conn = sqlite3.connect("parser.db")
    c = conn.cursor()
    c.execute("SELECT all_count, bad_number, bad_count, bad_year, bad_link FROM parsed_count WHERE id_parser = '" + str(id_parser) + "';")
    for all_count, bad_number, bad_count, bad_year, bad_link, in c.fetchall():
        pass
    new_all_count =int(all_count) + all_checked
    new_bad_number = int(bad_number) + bad_number1
    new_bad_count = int(bad_count) + bad_count1
    new_bad_year = int(bad_year) + bad_year1
    new_bad_link_checked = int(bad_link) + bad_link_checked1
    c.execute("UPDATE parsed_count SET all_count = '" + str(new_all_count) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_number = '" + str(new_bad_number) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_count = '" + str(new_bad_count) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_year = '" + str(new_bad_year) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_link = '" + str(new_bad_link_checked) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    conn.close()
    if thread == 15:
        conn = sqlite3.connect("parser.db")
        c = conn.cursor()
        c.execute("SELECT status FROM parser WHERE id_user = '" + str(chat_id) + "';")
        for status, in c.fetchall():
            if status == 0:
                c.execute("UPDATE parser SET  status = '" + '1' + "' WHERE id_user = '" + str(chat_id) + "';")
                conn.commit()
                sleep(5)
                mess = bot.send_message(chat_id, "<b>Ожидаем файл..</b>", parse_mode="html")
                sleep(2)
                wbs = openpyxl.Workbook()
                wb = wbs.active
                allvcf = open("Excel/"+str(chat_id) + '.vcf', 'w')
                wb.cell(row=1, column=1).value = "Название"
                wb.cell(row=1, column=2).value = "Ссылка"
                wb.cell(row=1, column=3).value = "Цена"
                wb.cell(row=1, column=4).value = "Имя"
                wb.cell(row=1, column=5).value = "Номер"
                wb.cell(row=1, column=6).value = "Количество объявлений"
                wb.cell(row=1, column=7).value = "Год создания"
                count_ads = 1
                c.execute("SELECT title, url, price, name, number, count_max_online, count_max_offline, year, img, gorod FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                for title, url, price, name, number, count_max_online, count_max_offline, year, img, gorod, in c.fetchall():
                    if count_ads == 1:
                        count_views = 0
                        c.execute("SELECT id_user FROM parsed_url WHERE url = '" + str(url) + "';")
                        for id_user, in c.fetchall():
                            count_views += 1
                        keys = types.InlineKeyboardMarkup()
                        keys_1 = types.InlineKeyboardButton(text="<<", callback_data="olx_checker_back_"+str(id_parser)+"/0")
                        keys_2 = types.InlineKeyboardButton(text="1 / **", callback_data="checker_nothing" + str(id_parser))
                        keys_3 = types.InlineKeyboardButton(text=">>", callback_data="olx_checker_next_" + str(id_parser) + "/2")
                        keys.add(keys_1, keys_2, keys_3)
                        if status_message == 'ON':
                            message_send = message_text.replace('[title]', title).replace('[link]', url)
                        else:
                            message_send = ''
                        bot.send_photo(chat_id, requests.get(img).content,
                                       caption=f"Название: <b>{title}</b>\n"
                                               f"Цена: <b>{price}</b>\n"
                                               f"Имя: <b>{name}</b>\n"
                                               f"Зарегистрирован: <b>{year}</b>\n"
                                               f"Местоположение: <b>{gorod}</b>\n"
                                               f"Номер: <b>+{number}</b>\n\n"
                                               f'Объявление: <a href="'+str(url)+'">*Клик*</a>\n'
                                               f'WhatsApp: <a href="https://api.whatsapp.com/send?phone={number}&text={message_send}">*Клик*</a>\n'
                                               f'Viber: <a href="https://msng.link/o/?{number}=vi">*Клик*</a>\n\n'
                                               f"Видело пользователей: <b>{count_views}</b>",
                                       parse_mode="html", reply_markup=keys)
                    try:
                        allvcf.write('BEGIN:VCARD' + "\n")
                        allvcf.write('N:' + str(title) + "\n")
                        allvcf.write('TEL;CELL:+' + str(number) + "\n")
                        allvcf.write("NOTE: " + str(url) + "\n")
                        allvcf.write('END:VCARD' + "\n")
                    except:
                        pass
                    info = [title, url, price, name, number, count_max_online, count_max_offline, year]
                    for x in info:
                        wb.cell(row=count_ads + 1, column=info.index(x) + 1).value = x
                    count_ads += 1
                wbs.save(f"Excel/{chat_id}.xlsx")
                allvcf.close()
                bot.delete_message(chat_id, mess.id)
                bot.send_document(chat_id, data=open(f"Excel/{chat_id}.xlsx", "rb"))
                bot.send_document(chat_id, data=open(f"Excel/{chat_id}.vcf", "rb"))
                c.execute("SELECT all_count, bad_number, bad_count, bad_year, bad_link FROM parsed_count WHERE id_parser = '" + str(id_parser) + "';")
                for all_count, bad_number, bad_count, bad_year, bad_link, in c.fetchall():
                    pass
                ddos = int(all_count) - int(bad_number) - int(bad_count) - int(bad_year) - int(bad_link)
                current_time = datetime.datetime.today()
                days = current_time - time_now
                time_all = round((int(days.days) * 86400 + int(days.seconds)) / 60, 1)
                text = f"<b>🦋 Парсинг окончен!</b>\n\n" \
                    f"- Количество проверенных: <b>{all_count}</b>\n" \
                    f"- Выдано: <b>{int(count_ads)-1}</b>\n" \
                    f"- Прошло: <b>{time_all} м</b>\n\n" \
                    f"<b>☠️ Количество отклонённых объявлений (Фильтры)</b>\n\n"  \
                    f"- Количество объявлений: <b>{bad_count}</b>\n" \
                    f"- Год создания: <b>{bad_year}</b>\n\n" \
                    f"- Отсутствует номер: <b>{bad_number}</b>\n" \
                    f"- Уже выдавалось вам: <b>{bad_link}</b>\n" \
                    f"- DDoS-защита: <b>{int(ddos)}</b>"
                if ddos > 100:
                    text += "\n\n<b>Активирована защита OLX\nПодождите 5м и запускайте парс!</b>"
                    print('Активирована защита OLX')
                elif count_need < int(need):
                    text += "\n\n<b>Смените категорию</b>"
                bot.send_message(chat_id, text, parse_mode="html")
                bot.send_message(chat_id, '🎉')
        conn.close()



if __name__ == '__main__':
    print('URL: ' + str(link) + '\nID: ' + str(id))
    dateFormatter = '%H/%d/%m/%Y'
    time_now = datetime.datetime.today()
    conn = sqlite3.connect("parser.db")
    c = conn.cursor()
    c.execute("INSERT INTO parsed_count VALUES(?,?,?,?,?,?)", (id_parser, '0', '0', '0', '0', '0'))
    conn.commit()
    conn.close()
    for i in range(1, 16):
        Process(target=parse_tr, args=(link, max_ads, need, min_year, status_message, message_text, id, country, i, id_parser, time_now)).start()
        sleep(0.5)
