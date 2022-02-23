# -*- coding: utf-8 -*-

import cfg
import telebot
from time import sleep
from sys import argv
import requests
from bs4 import BeautifulSoup
import json
from telebot import types
import sqlite3
from multiprocessing import Process
import openpyxl
import random
import datetime

bot = telebot.TeleBot(cfg.token)
script, link, need, status_message, message_text, id, id_parser, max_ads, min_year = argv






def parse_tr(link, need, status_message, message_text, chat_id, thread, id_parser, time_now, max_ads, min_year):
    checks = []
    main_link = link
    count_need = 0
    bad_number1 = 0
    bad_count1 = 0
    bad_year1 = 0
    bad_link_checked1 = 0
    all_checked = 0
    for start in range(int(thread), int(thread)+151, 10):
        try:
            if int(count_need) >= int(need):
                break
            if main_link[-1] == '/':
                check_page = requests.get(str(main_link) + 'cela-cr/cena-neomezena/nejnovejsi/'+str(start))
            else:
                check_page = requests.get(str(main_link) + '/cela-cr/cena-neomezena/nejnovejsi/' + str(start))
            checked = check_page.text
            soup = BeautifulSoup(checked, "html.parser")
            link_all = soup.find_all(class_="c-item__link")
            title_all = soup.find_all(class_="c-item__name-text")
            gorod_all = soup.find_all(class_="c-item__locality")
            for i in range(len(link_all)-1):
                try:
                    conn = sqlite3.connect("parser.db")
                    c = conn.cursor()
                    count_need = 0
                    c.execute("SELECT id_user FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                    for id_user, in c.fetchall():
                        count_need += 1
                    conn.close()
                    if int(count_need) >= int(need):
                        break
                    link_ad = link_all[i].get('href')
                    if link_ad in checks:
                        continue
                    try:
                        all_checked += 1
                        checks.append(link_ad)
                        check_page = requests.get(link_ad)
                        checked = check_page.text
                        number_1 = checked.find('tel:+') + 5
                        number_2 = checked.find('"', number_1)
                        number = checked[number_1:number_2]
                        try:
                            int(number)
                        except:
                            number = 'None'
                        if number == 'None':
                            bad_number1 += 1
                            continue
                        conn = sqlite3.connect("parser.db")
                        c = conn.cursor()
                        a = 0
                        count_views = 0
                        c.execute("SELECT id_user FROM parsed_url WHERE number = '" + str(number) + "';")
                        for id_user, in c.fetchall():
                            if int(id_user) == int(chat_id):
                                a = 1
                                break
                            count_views += 1
                        conn.close()
                        if a == 1:
                            bad_link_checked1 += 1
                            continue
                        prof_url_2 = link_ad.find('/', 23)
                        prof_url = str(link_ad[:prof_url_2])
                        p = requests.get(prof_url)
                        checked_prof = p.text
                        count_1 = checked_prof.find("<!-- -->(<!-- -->") + 17
                        count_2 = checked_prof.find("<", count_1)
                        count = checked_prof[count_1:count_2]
                        try:
                            if int(count) > int(max_ads):
                                bad_count1 += 1
                                continue
                        except:
                            count = 1
                        price_1 = checked.find('"price_original"') + 17
                        price_2 = checked.find(',', price_1)
                        try:
                            price = int(checked[price_1:price_2])
                            price = str(price) + " Kč"
                        except:
                            price = "Ошибка"
                        title = title_all[i].string
                        gorod = gorod_all[i + 1].string
                        img_1 = checked.find('content="//') + 11
                        img_2 = checked.find('"', img_1)
                        img = "http://" + checked[img_1:img_2]
                        link = link_ad
                        conn = sqlite3.connect("parser.db")
                        c = conn.cursor()
                        a = 0
                        c.execute("SELECT id_user FROM parsed_url WHERE url = '" + str(link_ad) + "';")
                        for id_user, in c.fetchall():
                            if int(id_user) == int(chat_id):
                                a = 1
                                break
                        conn.close()
                        if a == 1:
                            bad_link_checked1 += 1
                            continue
                        if status_message == 'ON':
                            message_send = message_text.replace('[title]', title).replace('[link]', link_ad)
                        else:
                            message_send = ''
                        name = link_ad[23:prof_url_2]
                        whatKey = types.InlineKeyboardMarkup()
                        key_last = types.InlineKeyboardButton(text="❌ Закрыть", callback_data="close")
                        whatKey.row(key_last)
                        conn = sqlite3.connect("parser.db")
                        c = conn.cursor()
                        year_born = ''
                        c.execute("INSERT INTO parsed_url VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", (chat_id, id_parser, link, title, name, price, number, count, "", year_born, img, gorod))
                        conn.commit()
                        conn.close()
                        bot.send_photo(chat_id, requests.get(img).content,
                                           caption=f"Название: <b>{title}</b>\n"
                                                   f"Цена: <b>{price}</b>\n"
                                                   f"Имя: <b>{name}</b>\n"
                                                   f"Местоположение: <b>{gorod}</b>\n"
                                                   f"Номер: <b>+{number}</b>\n\n"
                                                   f'Объявление: <a href="'+str(link_ad)+'">*Клик*</a>\n'
                                                   f'WhatsApp: <a href="https://api.whatsapp.com/send?phone={number}&text={message_send}">*Клик*</a>\n'
                                                   f'Viber: <a href="https://msng.link/o/?{number}=vi">*Клик*</a>\n\n'
                                                   f"Видело пользователей: <b>{count_views}</b>",
                                           parse_mode="html",
                                           reply_markup=whatKey)
                        conn = sqlite3.connect("parser.db")
                        c = conn.cursor()
                        count_need = 0
                        c.execute("SELECT id_user FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                        for id_user, in c.fetchall():
                            count_need += 1
                        conn.close()
                        if int(count_need) >= int(need):
                            break
                    except Exception as e:
                        continue
                except Exception as e:
                    continue
        except Exception as e:
            bot.send_message(chat_id, "⚠️ Ошибка")
            break
    conn = sqlite3.connect("parser.db")
    c = conn.cursor()
    c.execute("SELECT all_count, bad_number, bad_count, bad_year, bad_link FROM parsed_count WHERE id_parser = '" + str(id_parser) + "';")
    for all_count, bad_number, bad_count, bad_year, bad_link, in c.fetchall():
        pass
    new_all_count =int(all_count) + all_checked
    new_bad_number = int(bad_number) + bad_number1
    new_bad_count = int(bad_count) + bad_count1
    new_bad_year = int(bad_year) + bad_year1
    new_bad_link_checked = int(bad_link) + bad_link_checked1
    c.execute("UPDATE parsed_count SET all_count = '" + str(new_all_count) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_number = '" + str(new_bad_number) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_count = '" + str(new_bad_count) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_year = '" + str(new_bad_year) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    c.execute("UPDATE parsed_count SET bad_link = '" + str(new_bad_link_checked) + "' WHERE id_parser = '" + str(id_parser) + "';")
    conn.commit()
    conn.close()
    if thread == 15:
        conn = sqlite3.connect("parser.db")
        c = conn.cursor()
        c.execute("SELECT status FROM parser WHERE id_user = '" + str(chat_id) + "';")
        for status, in c.fetchall():
            if status == 0:
                c.execute("UPDATE parser SET  status = '" + '1' + "' WHERE id_user = '" + str(chat_id) + "';")
                conn.commit()
                sleep(5)
                mess = bot.send_message(chat_id, "<b>Ожидаем файл..</b>",parse_mode="html")
                sleep(2)
                wbs = openpyxl.Workbook()
                wb = wbs.active
                allvcf = open("Excel/"+str(chat_id) + '.vcf', 'w')
                wb.cell(row=1, column=1).value = "Название"
                wb.cell(row=1, column=2).value = "Ссылка"
                wb.cell(row=1, column=3).value = "Цена"
                wb.cell(row=1, column=4).value = "Имя"
                wb.cell(row=1, column=5).value = "Номер"
                count_ads = 1
                c.execute("SELECT title, url, price, name, number, count_max_online, count_max_offline, year, img, gorod FROM parsed_url WHERE id_parser = '" + str(id_parser) + "';")
                for title, url, price, name, number, count_max_online, count_max_offline, year, img, gorod, in c.fetchall():
                    if count_ads == 1:
                        count_views = 0
                        c.execute("SELECT id_user FROM parsed_url WHERE url = '" + str(url) + "';")
                        for id_user, in c.fetchall():
                            count_views += 1
                        keys = types.InlineKeyboardMarkup()
                        keys_1 = types.InlineKeyboardButton(text="<<", callback_data="sbazar_checker_back_"+str(id_parser)+"/0")
                        keys_2 = types.InlineKeyboardButton(text="1 / **", callback_data="checker_nothing" + str(id_parser))
                        keys_3 = types.InlineKeyboardButton(text=">>", callback_data="sbazar_checker_next_" + str(id_parser) + "/2")
                        keys.add(keys_1, keys_2, keys_3)
                        if status_message == 'ON':
                            message_send = message_text.replace('[title]', title).replace('[link]', url)
                        else:
                            message_send = ''
                        bot.send_photo(chat_id, requests.get(img).content,
                                       caption=f"Название: <b>{title}</b>\n"
                                               f"Цена: <b>{price}</b>\n"
                                               f"Имя: <b>{name}</b>\n"
                                               f"Местоположение: <b>{gorod}</b>\n"
                                               f"Номер: <b>+{number}</b>\n\n"
                                               f'Объявление: <a href="' + str(url) + '">*Клик*</a>\n'
                                               f'WhatsApp: <a href="https://api.whatsapp.com/send?phone={number}&text={message_send}">*Клик*</a>\n'
                                               f'Viber: <a href="https://msng.link/o/?{number}=vi">*Клик*</a>\n\n'
                                               f"Видело пользователей: <b>{count_views}</b>",
                                       parse_mode="html", reply_markup=keys)
                    info = [title, url, price, name, number]
                    for x in info:
                        wb.cell(row=count_ads + 1, column=info.index(x) + 1).value = x
                    try:
                        allvcf.write('BEGIN:VCARD' + "\n")
                        allvcf.write('N:' + str(title) + "\n")
                        allvcf.write('TEL;CELL:+' + str(number) + "\n")
                        allvcf.write("NOTE: " + str(url) + "\n")
                        allvcf.write('END:VCARD' + "\n")
                    except:
                        allvcf.write('N:' + 'Bugged' + "\n")
                        allvcf.write('TEL;CELL:+' + str(number) + "\n")
                        allvcf.write("NOTE: " + str(url) + "\n")
                        allvcf.write('END:VCARD' + "\n")
                    count_ads += 1
                wbs.save(f"Excel/{chat_id}.xlsx")
                allvcf.close()
                bot.delete_message(chat_id, mess.id)
                bot.send_document(chat_id, data=open(f"Excel/{chat_id}.xlsx", "rb"))
                bot.send_document(chat_id, data=open(f"Excel/{chat_id}.vcf", "rb"))
                c.execute("SELECT all_count, bad_number, bad_count, bad_year, bad_link FROM parsed_count WHERE id_parser = '" + str(id_parser) + "';")
                for all_count, bad_number, bad_count, bad_year, bad_link, in c.fetchall():
                    pass
                ddos = int(all_count) - int(bad_number) - int(bad_count) - int(bad_year) - int(bad_link)
                current_time = datetime.datetime.today()
                days = current_time - time_now
                time_all = round((int(days.days) * 86400 + int(days.seconds)) / 60, 1)
                text = f"<b>🦋 Парсинг окончен!</b>\n\n" \
                       f"- Количество проверенных: <b>{all_count}</b>\n" \
                       f"- Выдано: <b>{int(count_ads) - 1}</b>\n" \
                       f"- Прошло: <b>{time_all} м</b>\n\n" \
                       f"<b>☠️ Количество отклонённых объявлений (Фильтры)</b>\n\n" \
                       f"- Количество объявлений: <b>{bad_count}</b>\n" \
                       f"- Неверная цена: <b>{int(ddos)}</b>\n" \
                       f"- Год создания: <b>{bad_year}</b>\n\n" \
                       f"- Отсутствует номер: <b>{bad_number}</b>\n" \
                       f"- Уже выдавалось вам: <b>{bad_link}</b>\n"
                if count_need < int(need):
                    text += "\n\n<b>Смените категорию</b>"
                bot.send_message(chat_id, text, parse_mode="html")
                bot.send_message(chat_id, '🎉')
        conn.close()



if __name__ == '__main__':
    print('URL: ' + str(link) + '\nID: ' + str(id))
    dateFormatter = '%H/%d/%m/%Y'
    time_now = datetime.datetime.today()
    conn = sqlite3.connect("parser.db")
    c = conn.cursor()
    c.execute("INSERT INTO parsed_count VALUES(?,?,?,?,?,?)", (id_parser, '0', '0', '0', '0', '0'))
    conn.commit()
    conn.close()
    for i in range(1, 16):
        Process(target=parse_tr, args=(link, need, status_message, message_text, id, i, id_parser, time_now, max_ads, min_year)).start()
        sleep(0.5)
